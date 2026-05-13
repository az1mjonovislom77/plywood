import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class AcceptanceExportService:
    HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")
    GROUP_FILL = PatternFill("solid", fgColor="F8FAFC")
    THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))

    @staticmethod
    def build_supplier_excel(queryset, supplier_name, date):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Acceptance"
        title = f"{supplier_name} - {date.strftime('%d.%m.%Y')}"
        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["Mahsulot", "Miqdor", "Kelish narxi", "Investitsiya"]
        ws.append([])
        ws.append(headers)

        for cell in ws[3]:
            cell.font = Font(bold=True)

        total_quantity = 0
        total_investment = 0

        for obj in queryset:
            investment = obj.arrival_price * obj.count
            ws.append([
                obj.product.name,
                float(obj.count),
                float(obj.arrival_price),
                float(investment),
            ])

            total_quantity += obj.count
            total_investment += investment

        for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = '#,##0'

        ws.append([])
        ws.append([
            "JAMI",
            float(total_quantity),
            "",
            float(total_investment)
        ])

        last_row = ws.max_row

        for cell in ws[last_row]:
            cell.font = Font(bold=True)
            if cell.column in [2, 4]:
                cell.number_format = '#,##0'

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 22

        return wb

    @classmethod
    def build_analytics_excel(cls, data, from_date=None, to_date=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics"

        title = "Qabul qilish tahlili"
        if from_date and to_date:
            title = f"{title} ({from_date.strftime('%d.%m.%Y')} - {to_date.strftime('%d.%m.%Y')})"
        elif from_date:
            title = f"{title} ({from_date.strftime('%d.%m.%Y')} dan)"
        elif to_date:
            title = f"{title} ({to_date.strftime('%d.%m.%Y')} gacha)"

        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["Sana", "Yetkazib beruvchi", "Miqdor", "Investitsiya"])
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER

        total_quantity = 0
        total_investment = 0

        for group in data:
            suppliers = group["suppliers"]
            date = group["date"]
            ws.append([
                date.strftime("%d.%m.%Y"),
                f"{len(suppliers)} ta yetkazib beruvchi",
                "",
                "",
            ])
            group_row = ws.max_row
            for cell in ws[group_row]:
                cell.font = Font(bold=True)
                cell.fill = cls.GROUP_FILL
                cell.border = cls.THIN_BORDER

            for supplier in suppliers:
                quantity = supplier["total_quantity"] or 0
                investment = supplier["total_investment"] or 0
                ws.append([
                    "",
                    supplier["supplier_name"],
                    float(quantity),
                    float(investment),
                ])
                row = ws.max_row
                ws.cell(row=row, column=3).number_format = '#,##0.### "dona"'
                ws.cell(row=row, column=4).number_format = '#,##0.## "UZS"'
                for cell in ws[row]:
                    cell.border = cls.THIN_BORDER

                total_quantity += quantity
                total_investment += investment

        ws.append([])
        ws.append(["JAMI", "", float(total_quantity), float(total_investment)])
        total_row = ws.max_row
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER
        ws.cell(row=total_row, column=3).number_format = '#,##0.### "dona"'
        ws.cell(row=total_row, column=4).number_format = '#,##0.## "UZS"'
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 24

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            row[2].alignment = Alignment(horizontal="right")
            row[3].alignment = Alignment(horizontal="right")

        return wb
