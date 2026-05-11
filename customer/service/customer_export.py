from decimal import Decimal
from io import BytesIO
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from customer.models import Customer, BalanceHistory
from order.models import Order


class CustomerStatementService:
    @classmethod
    def _parse_bounds(cls, date_from, date_to):
        today = timezone.localdate()
        start_date = parse_date(date_from) if date_from else today
        end_date = parse_date(date_to) if date_to else today

        if not start_date or not end_date:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        if end_date < start_date:
            raise ValueError("to date must be greater than or equal to from date")

        start_dt = timezone.make_aware(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_dt = timezone.make_aware(
            timezone.datetime.combine(end_date + timezone.timedelta(days=1), timezone.datetime.min.time())
        )
        return start_date, end_date, start_dt, end_dt

    @classmethod
    def _opening_balance(cls, customer_id, start_dt):
        previous_orders = (
            Order.objects
            .filter(customer_id=customer_id, created_at__lt=start_dt)
            .exclude(order_status=Order.OrderStatus.CANCEL)
            .aggregate(
                total=Coalesce(Sum("total_price"), Value(Decimal("0")), output_field=DecimalField()),
                paid=Coalesce(Sum("covered_amount"), Value(Decimal("0")), output_field=DecimalField()),
            )
        )
        manual_payments = BalanceHistory.objects.filter(
            customer_id=customer_id,
            type=BalanceHistory.Type.PAYMENT,
            created_at__lt=start_dt
        ).aggregate(total_paid=Coalesce(Sum("amount"), Value(Decimal("0"))))

        total_paid = previous_orders["paid"] + manual_payments["total_paid"]
        total_ordered = previous_orders["total"]
        return total_paid - total_ordered

    @classmethod
    def build_statement(cls, customer_id, date_from=None, date_to=None):
        customer = Customer.objects.get(pk=customer_id)
        start_date, end_date, start_dt, end_dt = cls._parse_bounds(date_from, date_to)
        running_balance = cls._opening_balance(customer_id, start_dt)

        orders = (
            Order.objects
            .filter(customer_id=customer_id, created_at__gte=start_dt, created_at__lt=end_dt)
            .exclude(order_status=Order.OrderStatus.CANCEL)
            .order_by("created_at", "id")
        )

        payments = (
            BalanceHistory.objects
            .filter(
                customer_id=customer_id,
                type=BalanceHistory.Type.PAYMENT,
                created_at__gte=start_dt,
                created_at__lt=end_dt,
            ).order_by("created_at", "id"))

        events = []

        for order in orders:
            events.append(
                (order.created_at, 0,
                 [{
                     "date": order.created_at.date(),
                     "registrator": f"Продажа товара {order.id:09d} от {order.created_at:%d.%m.%Y %H:%M:%S}",
                     "payment_type": order.get_payment_method_display(),
                     "purpose": None,
                     "product": f"Order {order.id}",
                     "income_qty": None,
                     "income_amount": None,
                     "expense_qty": 1,
                     "expense_amount": order.total_price,
                 }
                 ],
                 )
            )

        for payment in payments:
            events.append(
                (payment.created_at, 1,
                 [{
                     "date": payment.created_at.date(),
                     "registrator": f"Оплата клиента {payment.id:09d} от {payment.created_at:%d.%m.%Y %H:%M:%S}",
                     "payment_type": "Наличная",
                     "purpose": None,
                     "product": None,
                     "income_qty": None,
                     "income_amount": payment.amount,
                     "expense_qty": None,
                     "expense_amount": None,
                 }])
            )

        events.sort(key=lambda event: (event[0], event[1]))

        rows = []
        total_income_amount = Decimal("0")
        total_expense_amount = Decimal("0")
        no = 1

        for _, _, event_rows in events:
            for row in event_rows:
                if row["income_amount"]:
                    running_balance += row["income_amount"]
                    total_income_amount += row["income_amount"]

                if row["expense_amount"]:
                    running_balance -= row["expense_amount"]
                    total_expense_amount += row["expense_amount"]

                row["no"] = no
                row["balance"] = running_balance
                rows.append(row)
                no += 1

        return {
            "supplier": {
                "id": customer.id,
                "full_name": customer.full_name,
            },
            "from": start_date,
            "to": end_date,
            "opening_balance": cls._opening_balance(customer_id, start_dt),
            "rows": rows,
            "totals": {
                "income_amount": total_income_amount,
                "expense_amount": total_expense_amount,
                "closing_balance": running_balance,
            },
        }

    @staticmethod
    def _fmt_number(value):
        if value is None:
            return None
        if isinstance(value, Decimal) and value == value.to_integral_value():
            return int(value)
        return float(value)

    @staticmethod
    def _apply_table_style(ws, data_start_row, data_end_row):
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        base_font = Font(name="Arial", size=8)
        bold_font = Font(name="Arial", size=8, bold=True)
        negative_font = Font(name="Arial", size=8, color="00C00000")

        widths = {
            "A": 13, "B": 5.28515625, "C": 13, "D": 35.42578125,
            "E": 15.7109375, "F": 14.7109375, "G": 26,
            "H": 13, "I": 13, "J": 13, "K": 13, "L": 13,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("C1:D1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("C3:E3")
        ws.merge_cells("H4:I4")
        ws.merge_cells("J4:K4")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:C5")
        ws.merge_cells("D4:D5")
        ws.merge_cells("E4:E5")
        ws.merge_cells("F4:F5")
        ws.merge_cells("G4:G5")
        ws.merge_cells("L4:L5")
        ws.merge_cells(f"B{data_end_row}:E{data_end_row}")

        for row in range(4, 6):
            for col in range(2, 13):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center

        for row in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[row].height = 67.5
            for col in range(2, 13):
                cell = ws.cell(row=row, column=col)
                cell.font = base_font
                cell.border = border
                if col in [2, 3, 8, 9, 10, 11, 12]:
                    cell.alignment = center if col in [2, 3, 8, 10] else right
                elif col in [4, 7]:
                    cell.alignment = left
                else:
                    cell.alignment = center

                if col in [9, 11, 12] and cell.value is not None:
                    cell.number_format = "#,##0"
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = negative_font

        ws["C1"].font = bold_font
        ws["C2"].font = bold_font
        ws["C3"].font = bold_font
        ws["K3"].font = bold_font
        ws["L3"].font = bold_font if (ws["L3"].value or 0) >= 0 else Font(name="Arial", size=8, bold=True,
                                                                          color="00C00000")
        ws["L3"].number_format = "#,##0"
        ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
        ws["K3"].alignment = center
        ws["L3"].alignment = right
        total_row = data_end_row
        ws.cell(row=total_row, column=2).font = bold_font
        ws.cell(row=total_row, column=2).alignment = right
        ws.cell(row=total_row, column=9).font = bold_font
        ws.cell(row=total_row, column=11).font = bold_font
        ws.cell(row=total_row, column=9).number_format = "#,##0"
        ws.cell(row=total_row, column=11).number_format = "#,##0"

    @classmethod
    def build_statement_excel(cls, customer_id, date_from=None, date_to=None):
        data = cls.build_statement(customer_id=customer_id, date_from=date_from, date_to=date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["C1"] = f"{data['from'].strftime('%d.%m.%Y')} 0:00:00"
        ws["C2"] = f"{data['to'].strftime('%d.%m.%Y')} 23:59:59"
        ws["C3"] = data["supplier"]["full_name"]
        ws["K3"] = "Остаток"
        ws["L3"] = cls._fmt_number(data["opening_balance"])
        ws["B4"] = "№"
        ws["C4"] = "Дата"
        ws["D4"] = "Регистратор"
        ws["E4"] = "Вид оплаты"
        ws["F4"] = "ТуловМаксади"
        ws["G4"] = "Товар"
        ws["H4"] = "Приход"
        ws["J4"] = "Расход"
        ws["L4"] = "Остаток"
        ws["H5"] = "Кол"
        ws["I5"] = "Сумма"
        ws["J5"] = "Кол"
        ws["K5"] = "Сумма"

        row_idx = 6
        data_start_row = row_idx

        for row in data["rows"]:
            ws.cell(row=row_idx, column=2, value=row["no"])
            ws.cell(row=row_idx, column=3, value=row["date"].strftime("%d.%m.%Y"))
            ws.cell(row=row_idx, column=4, value=row["registrator"])
            ws.cell(row=row_idx, column=5, value=row["payment_type"])
            ws.cell(row=row_idx, column=6, value=row["purpose"])
            ws.cell(row=row_idx, column=7, value=row["product"])
            ws.cell(row=row_idx, column=8, value=cls._fmt_number(row["income_qty"]))
            ws.cell(row=row_idx, column=9, value=cls._fmt_number(row["income_amount"]))
            ws.cell(row=row_idx, column=10, value=cls._fmt_number(row["expense_qty"]))
            ws.cell(row=row_idx, column=11, value=cls._fmt_number(row["expense_amount"]))
            ws.cell(row=row_idx, column=12, value=cls._fmt_number(row["balance"]))
            row_idx += 1

        ws.cell(row=row_idx, column=2, value="Жами:")
        ws.cell(row=row_idx, column=9, value=cls._fmt_number(data["totals"]["income_amount"]))
        ws.cell(row=row_idx, column=11, value=cls._fmt_number(data["totals"]["expense_amount"]))
        cls._apply_table_style(ws, data_start_row=data_start_row, data_end_row=row_idx)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class SalesStatementService:
    @classmethod
    def _parse_bounds(cls, date_from, date_to):
        today = timezone.localdate()
        start_date = parse_date(date_from) if date_from else today
        end_date = parse_date(date_to) if date_to else today

        if not start_date or not end_date:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        if end_date < start_date:
            raise ValueError("to date must be greater than or equal to from date")

        start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(
            timezone.datetime.combine(end_date + timezone.timedelta(days=1), timezone.datetime.min.time())
        )
        return start_date, end_date, start_dt, end_dt

    @classmethod
    def build_statement(cls, customer_id=None, date_from=None, date_to=None):
        start_date, end_date, start_dt, end_dt = cls._parse_bounds(date_from, date_to)

        qs = (
            Order.objects
            .filter(created_at__gte=start_dt, created_at__lt=end_dt)
            .exclude(order_status=Order.OrderStatus.CANCEL)
            .select_related("customer")
            .order_by("created_at", "id")
        )

        if customer_id:
            qs = qs.filter(customer_id=customer_id)

        rows = []
        total_amount = Decimal("0")
        no = 1

        for item in qs:
            total_amount += item.total_price
            rows.append(
                {
                    "no": no,
                    "date": item.created_at.date(),
                    "order_no": item.id,
                    "doc_type": "Продажа товара",
                    "supplier": item.customer.full_name if item.customer else "Аноним",
                    "amount": item.total_price,
                }
            )
            no += 1

        customer_name = "Все клиенты"
        if customer_id:
            customer = Customer.objects.get(pk=customer_id)
            customer_name = customer.full_name

        return {
            "supplier_name": customer_name,
            "from": start_date,
            "to": end_date,
            "rows": rows,
            "total_amount": total_amount,
        }

    @staticmethod
    def _fmt_number(value):
        if value is None:
            return None
        if isinstance(value, Decimal) and value == value.to_integral_value():
            return int(value)
        return float(value)

    @staticmethod
    def _apply_table_style(ws, data_start_row, data_end_row):
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        base_font = Font(name="Arial", size=8)
        bold_font = Font(name="Arial", size=8, bold=True)

        widths = {"A": 13, "B": 6, "C": 14, "D": 18, "E": 22, "F": 28, "G": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("C1:D1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("C3:E3")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:C5")
        ws.merge_cells("D4:D5")
        ws.merge_cells("E4:E5")
        ws.merge_cells("F4:F5")
        ws.merge_cells("G4:G5")
        ws.merge_cells(f"B{data_end_row}:F{data_end_row}")

        for row in range(4, 6):
            for col in range(2, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center

        for row in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[row].height = 42
            for col in range(2, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = base_font
                cell.border = border

                if col in [2, 3, 4]:
                    cell.alignment = center
                elif col in [5, 6]:
                    cell.alignment = left
                else:
                    cell.alignment = right
                    if cell.value is not None:
                        cell.number_format = "#,##0"

        ws["C1"].font = bold_font
        ws["C2"].font = bold_font
        ws["C3"].font = bold_font
        ws["C3"].alignment = left
        total_row = data_end_row
        ws.cell(row=total_row, column=2).font = bold_font
        ws.cell(row=total_row, column=2).alignment = right
        ws.cell(row=total_row, column=7).font = bold_font
        ws.cell(row=total_row, column=7).number_format = "#,##0"

    @classmethod
    def build_statement_excel(cls, customer_id=None, date_from=None, date_to=None):
        data = cls.build_statement(customer_id=customer_id, date_from=date_from, date_to=date_to)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["C1"] = f"{data['from'].strftime('%d.%m.%Y')} 0:00:00"
        ws["C2"] = f"{data['to'].strftime('%d.%m.%Y')} 23:59:59"
        ws["C3"] = data["supplier_name"]
        ws["B4"] = "№"
        ws["C4"] = "Дата"
        ws["D4"] = "Номер"
        ws["E4"] = "Тип документа"
        ws["F4"] = "Клиент"
        ws["G4"] = "Сумма"

        row_idx = 6
        data_start_row = row_idx

        for row in data["rows"]:
            ws.cell(row=row_idx, column=2, value=row["no"])
            ws.cell(row=row_idx, column=3, value=row["date"].strftime("%d.%m.%Y"))
            ws.cell(row=row_idx, column=4, value=row["order_no"])
            ws.cell(row=row_idx, column=5, value=row["doc_type"])
            ws.cell(row=row_idx, column=6, value=row["supplier"])
            ws.cell(row=row_idx, column=7, value=cls._fmt_number(row["amount"]))
            row_idx += 1

        ws.cell(row=row_idx, column=2, value="Жами:")
        ws.cell(row=row_idx, column=7, value=cls._fmt_number(data["total_amount"]))
        cls._apply_table_style(ws, data_start_row=data_start_row, data_end_row=row_idx)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
