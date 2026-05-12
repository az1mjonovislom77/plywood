from io import BytesIO
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from customer.models import Customer
from customer.service.customer_balance import CustomerBalanceService


class CustomerDebtExcelService:

    @staticmethod
    def _style(ws, last_row):

        thin = Side(style='thin', color='000000')

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        header_font = Font(
            bold=True,
            size=12
        )

        body_font = Font(size=12)

        center = Alignment(
            horizontal='center',
            vertical='center'
        )

        right = Alignment(
            horizontal='right',
            vertical='center'
        )

        left = Alignment(
            horizontal='left',
            vertical='center'
        )

        widths = {
            1: 8,
            2: 45,
            3: 20,
            4: 20,
        }

        for col, width in widths.items():
            ws.column_dimensions[
                get_column_letter(col)
            ].width = width

        for row in ws.iter_rows(
                min_row=1,
                max_row=last_row,
                min_col=1,
                max_col=4
        ):
            for cell in row:
                cell.border = border
                cell.font = body_font

        for row in [1, 3]:
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = center

        for row in range(5, last_row + 1):
            ws.cell(
                row=row,
                column=1
            ).alignment = center

            ws.cell(
                row=row,
                column=2
            ).alignment = left

            ws.cell(
                row=row,
                column=3
            ).alignment = right

            ws.cell(
                row=row,
                column=4
            ).alignment = right

            ws.cell(
                row=row,
                column=3
            ).number_format = '#,##0.00'

            ws.cell(
                row=row,
                column=4
            ).number_format = '#,##0.00'

    @classmethod
    def build(cls, date_from='2024-01-01', date_to=None):

        today = timezone.localdate()

        start_date = (parse_date(date_from)
                      if date_from else parse_date('2024-01-01'))

        end_date = (
            parse_date(date_to)
            if date_to else today
        )

        wb = Workbook()

        ws = wb.active

        ws.merge_cells('A1:D1')

        ws['A1'] = (
            f'{start_date.strftime("%d.%m.%Y")}'
            f' - '
            f'{end_date.strftime("%d.%m.%Y")}'
        )

        ws['A3'] = '№'
        ws['B3'] = 'Mijozlar'
        ws['C3'] = 'Ortiqcha to`lov qilganlar'
        ws['D3'] = 'Qarzdorlar'

        row = 5

        total_dt = Decimal("0")
        total_kt = Decimal("0")

        customers = Customer.objects.all().order_by(
            "full_name"
        )

        for index, customer in enumerate(customers, start=1):

            debt = CustomerBalanceService.calculate_customer_debt(
                customer=customer, date_from=start_date, date_to=end_date)

            debt = Decimal(str(debt or 0))

            dt = None
            kt = None

            if debt < 0:
                dt = abs(debt)
                total_dt += abs(debt)

            elif debt > 0:
                kt = debt
                total_kt += debt

            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=customer.full_name)
            ws.cell(row=row, column=3, value=float(kt) if kt else None)
            ws.cell(row=row, column=4, value=float(dt) if dt else None)

            row += 1

        ws.cell(row=row, column=2, value='Жами:')
        ws.cell(row=row, column=3, value=float(total_kt))
        ws.cell(row=row, column=4, value=float(total_dt))

        cls._style(ws, row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @classmethod
    def response(cls, request):
        date_from = request.GET.get("from")
        date_to = request.GET.get("to")

        # Fallback to default if date_from is empty string from request
        if not date_from:
            date_from = '2024-01-01'

        output = cls.build(date_from=date_from, date_to=date_to)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )

        response['Content-Disposition'] = (
            'attachment; filename="customer_debt_report.xlsx"'
        )

        return response
