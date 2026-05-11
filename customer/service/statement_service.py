from decimal import Decimal
from io import BytesIO
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from customer.models import Customer, BalanceHistory
from customer.service.customer_balance import CustomerBalanceService
from order.models import Order


class CustomerStatementService:
    PAYMENT_LABELS = {
        "cash": "Наличная",
        "card": "Пластик карта",
        "nasiya": "Насия",
    }

    @staticmethod
    def _service_total(service):
        total = service.calculate_price()

        if service.discount > 0:
            if service.discount_type == service.DiscountType.PERCENTAGE:
                total -= total * (service.discount / Decimal("100"))
            else:
                total -= service.discount

        return max(total, Decimal("0"))

    @classmethod
    def _parse_bounds(cls, date_from, date_to):
        today = timezone.localdate()
        start_date = parse_date(date_from) if date_from else today
        end_date = parse_date(date_to) if date_to else today

        if not start_date or not end_date:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        if end_date < start_date:
            raise ValueError("to date must be greater than or equal to from date")

        start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(
            timezone.datetime.combine(end_date + timezone.timedelta(days=1), timezone.datetime.min.time())
        )
        return start_date, end_date, start_dt, end_dt

    @classmethod
    def _opening_balance(cls, customer_id, start_dt):
        from order.models import Banding, Cutting

        active_orders = Order.objects.filter(
            customer_id=customer_id, created_at__lt=start_dt
        ).exclude(order_status=Order.OrderStatus.CANCEL)

        cancelled_orders = Order.objects.filter(
            customer_id=customer_id, created_at__lt=start_dt, order_status=Order.OrderStatus.CANCEL
        )

        standalone_bandings = Banding.objects.filter(
            customer_id=customer_id, created_at__lt=start_dt, orders__isnull=True, order_items__isnull=True
        )

        standalone_cuttings = Cutting.objects.filter(
            customer_id=customer_id, created_at__lt=start_dt, orders__isnull=True, order_items__isnull=True
        )

        manual_payments = BalanceHistory.objects.filter(
            customer_id=customer_id, type=BalanceHistory.Type.PAYMENT, created_at__lt=start_dt
        )

        total_payments = sum(p.amount for p in manual_payments)
        total_cancelled_covered = sum(o.covered_amount for o in cancelled_orders)
        active_order_balance = sum(o.covered_amount - o.total_price for o in active_orders)
        banding_balance = sum(b.covered_amount - cls._service_total(b) for b in standalone_bandings)
        cutting_balance = sum(c.covered_amount - cls._service_total(c) for c in standalone_cuttings)

        return total_payments + total_cancelled_covered + active_order_balance + banding_balance + cutting_balance

    @classmethod
    def build_statement(cls, customer_id, date_from=None, date_to=None):
        customer = Customer.objects.get(pk=customer_id)
        start_date, end_date, start_dt, end_dt = cls._parse_bounds(date_from, date_to)
        running_balance = cls._opening_balance(customer_id, start_dt)

        orders = (
            Order.objects
            .filter(customer_id=customer_id, created_at__gte=start_dt, created_at__lt=end_dt)
            .exclude(order_status=Order.OrderStatus.CANCEL)
            .select_related("banding__thickness", "cutting")
            .prefetch_related("items__product", "items__banding__thickness", "items__cutting")
            .order_by("created_at", "id")
        )

        from order.models import Banding, Cutting

        standalone_bandings = (
            Banding.objects
            .filter(customer_id=customer_id, created_at__gte=start_dt, created_at__lt=end_dt, orders__isnull=True,
                    order_items__isnull=True)
            .select_related("thickness")
            .order_by("created_at", "id")
        )

        standalone_cuttings = (
            Cutting.objects
            .filter(customer_id=customer_id, created_at__gte=start_dt, created_at__lt=end_dt, orders__isnull=True,
                    order_items__isnull=True)
            .order_by("created_at", "id")
        )

        manual_payments = (
            BalanceHistory.objects
            .filter(
                customer_id=customer_id,
                type=BalanceHistory.Type.PAYMENT,
                created_at__gte=start_dt,
                created_at__lt=end_dt,
            ).order_by("created_at", "id")
        )

        events = []

        for order in orders:
            rows = []
            registrator = f"Продажа товара {order.id:09d} от {order.created_at:%d.%m.%Y %H:%M:%S}"

            for item in order.items.all():
                product_amount = item.price * item.quantity
                rows.append(
                    {
                        "date": order.created_at.date(),
                        "registrator": registrator,
                        "payment_type": None,
                        "purpose": None,
                        "product": item.product.name,
                        "income_qty": None,
                        "income_amount": None,
                        "expense_qty": item.quantity,
                        "expense_amount": product_amount,
                    }
                )

                if item.cutting:
                    rows.append(
                        {
                            "date": order.created_at.date(),
                            "registrator": registrator,
                            "payment_type": None,
                            "purpose": None,
                            "product": "Хизмат (Распил)",
                            "income_qty": None,
                            "income_amount": None,
                            "expense_qty": item.cutting.count,
                            "expense_amount": cls._service_total(item.cutting),
                        }
                    )

                if item.banding:
                    thickness = item.banding.thickness.text if item.banding.thickness else ""
                    rows.append(
                        {
                            "date": order.created_at.date(),
                            "registrator": registrator,
                            "payment_type": None,
                            "purpose": None,
                            "product": f"Хизмат (Кромка) {thickness}".strip(),
                            "income_qty": None,
                            "income_amount": None,
                            "expense_qty": item.banding.length,
                            "expense_amount": cls._service_total(item.banding),
                        }
                    )

            if order.cutting:
                rows.append(
                    {
                        "date": order.created_at.date(),
                        "registrator": registrator,
                        "payment_type": None,
                        "purpose": None,
                        "product": "Хизмат (Распил)",
                        "income_qty": None,
                        "income_amount": None,
                        "expense_qty": order.cutting.count,
                        "expense_amount": cls._service_total(order.cutting),
                    }
                )

            if order.banding:
                thickness = order.banding.thickness.text if order.banding.thickness else ""
                rows.append(
                    {
                        "date": order.created_at.date(),
                        "registrator": registrator,
                        "payment_type": None,
                        "purpose": None,
                        "product": f"Хизмат (Кромка) {thickness}".strip(),
                        "income_qty": None,
                        "income_amount": None,
                        "expense_qty": order.banding.length,
                        "expense_amount": cls._service_total(order.banding),
                    }
                )

            total_items_services = sum(r["expense_amount"] for r in rows if r["expense_amount"])
            order_discount = total_items_services - order.total_price

            if order_discount > 0:
                rows.append(
                    {
                        "date": order.created_at.date(),
                        "registrator": registrator,
                        "payment_type": None,
                        "purpose": None,
                        "product": "Chegirma",
                        "income_qty": None,
                        "income_amount": None,
                        "expense_qty": None,
                        "expense_amount": -order_discount,
                    })

            if order.covered_amount > 0:
                rows.append(
                    {
                        "date": order.created_at.date(),
                        "registrator": registrator,
                        "payment_type": cls.PAYMENT_LABELS.get(order.payment_method, order.payment_method),
                        "purpose": None,
                        "product": None,
                        "income_qty": None,
                        "income_amount": order.covered_amount,
                        "expense_qty": None,
                        "expense_amount": None,
                    }
                )

            events.append((order.created_at, 0, rows))

        for banding in standalone_bandings:
            rows = []
            thickness = banding.thickness.text if banding.thickness else ""
            registrator = f"Хизмат {banding.id:09d} от {banding.created_at:%d.%m.%Y %H:%M:%S}"
            rows.append({
                "date": banding.created_at.date(),
                "registrator": registrator,
                "payment_type": None,
                "purpose": None,
                "product": f"Хизмат (Кромка) {thickness}".strip(),
                "income_qty": None,
                "income_amount": None,
                "expense_qty": banding.length,
                "expense_amount": cls._service_total(banding),
            })
            if banding.covered_amount > 0:
                rows.append({
                    "date": banding.created_at.date(),
                    "registrator": registrator,
                    "payment_type": cls.PAYMENT_LABELS.get(banding.payment_method, banding.payment_method),
                    "purpose": None,
                    "product": None,
                    "income_qty": None,
                    "income_amount": banding.covered_amount,
                    "expense_qty": None,
                    "expense_amount": None,
                })
            events.append((banding.created_at, 0, rows))

        for cutting in standalone_cuttings:
            rows = []
            registrator = f"Хизмат {cutting.id:09d} от {cutting.created_at:%d.%m.%Y %H:%M:%S}"
            rows.append({
                "date": cutting.created_at.date(),
                "registrator": registrator,
                "payment_type": None,
                "purpose": None,
                "product": "Хизмат (Распил)",
                "income_qty": None,
                "income_amount": None,
                "expense_qty": cutting.count,
                "expense_amount": cls._service_total(cutting),
            })
            if cutting.covered_amount > 0:
                rows.append({
                    "date": cutting.created_at.date(),
                    "registrator": registrator,
                    "payment_type": cls.PAYMENT_LABELS.get(cutting.payment_method, cutting.payment_method),
                    "purpose": None,
                    "product": None,
                    "income_qty": None,
                    "income_amount": cutting.covered_amount,
                    "expense_qty": None,
                    "expense_amount": None,
                })
            events.append((cutting.created_at, 0, rows))

        for payment in manual_payments:
            events.append(
                (payment.created_at, 1, [{
                    "date": payment.created_at.date(),
                    "registrator": f"Касса {payment.id} от {payment.created_at:%d.%m.%Y %H:%M:%S}",
                    "payment_type": "Наличная",
                    "purpose": None,
                    "product": None,
                    "income_qty": None,
                    "income_amount": payment.amount,
                    "expense_qty": None,
                    "expense_amount": None
                }]))

        cancelled_orders_current_period = (
            Order.objects
            .filter(customer_id=customer_id, created_at__gte=start_dt, created_at__lt=end_dt,
                    order_status=Order.OrderStatus.CANCEL)
            .order_by("created_at", "id")
        )

        for co in cancelled_orders_current_period:
            if co.covered_amount > 0:
                registrator = f"Бекор қилинган буюртма тўлови {co.id:09d} от {co.created_at:%d.%m.%Y %H:%M:%S}"
                rows = [
                    {
                        "date": co.created_at.date(),
                        "registrator": registrator,
                        "payment_type": cls.PAYMENT_LABELS.get(co.payment_method, co.payment_method),
                        "purpose": None,
                        "product": "Возврат / Отмененный заказ",
                        "income_qty": None,
                        "income_amount": co.covered_amount,
                        "expense_qty": None,
                        "expense_amount": None,
                    }
                ]
                events.append((co.created_at, 0, rows))

        events.sort(key=lambda event: (event[0], event[1]))
        rows = []
        total_income_amount = Decimal("0")
        total_expense_amount = Decimal("0")
        no = 1

        for _, _, event_rows in events:
            for row in event_rows:
                if row["income_amount"]:
                    running_balance += row["income_amount"]
                    total_income_amount += row["income_amount"]

                if row["expense_amount"]:
                    running_balance -= row["expense_amount"]
                    total_expense_amount += row["expense_amount"]

                row["no"] = no
                row["balance"] = running_balance
                rows.append(row)
                no += 1

        balance_stats = CustomerBalanceService.calculate(customer_id)

        return {
            "customer": {
                "id": customer.id,
                "full_name": customer.full_name,
            },
            "from": start_date,
            "to": end_date,
            "opening_balance": cls._opening_balance(customer_id, start_dt),
            "rows": rows,
            "totals": {
                "income_amount": total_income_amount,
                "expense_amount": total_expense_amount,
                "closing_balance": balance_stats["remaining_debt"],
            },
        }

    @staticmethod
    def _fmt_number(value):
        if value is None:
            return None
        if isinstance(value, Decimal) and value == value.to_integral_value():
            return int(value)
        return float(value)

    @staticmethod
    def _apply_table_style(ws, data_start_row, data_end_row):
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        base_font = Font(name="Arial", size=8)
        bold_font = Font(name="Arial", size=8, bold=True)
        negative_font = Font(name="Arial", size=8, color="00C00000")

        widths = {
            "A": 13, "B": 5.28515625, "C": 13, "D": 35.42578125, "E": 15.7109375, "F": 14.7109375, "G": 26,
            "H": 13, "I": 13, "J": 13, "K": 13, "L": 13}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("C1:D1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("C3:E3")
        ws.merge_cells("H4:I4")
        ws.merge_cells("J4:K4")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:C5")
        ws.merge_cells("D4:D5")
        ws.merge_cells("E4:E5")
        ws.merge_cells("F4:F5")
        ws.merge_cells("G4:G5")
        ws.merge_cells("L4:L5")
        ws.merge_cells(f"B{data_end_row}:E{data_end_row}")

        for row in range(4, 6):
            for col in range(2, 13):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center

        for row in range(data_start_row, data_end_row + 1):
            ws.row_dimensions[row].height = 67.5
            for col in range(2, 13):
                cell = ws.cell(row=row, column=col)
                cell.font = base_font
                cell.border = border
                if col in [2, 3, 8, 9, 10, 11, 12]:
                    cell.alignment = center if col in [2, 3, 8, 10] else right
                elif col in [4, 7]:
                    cell.alignment = left
                else:
                    cell.alignment = center

                if col in [9, 11, 12] and cell.value is not None:
                    cell.number_format = "#,##0"
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = negative_font

        ws["C1"].font = bold_font
        ws["C2"].font = bold_font
        ws["C3"].font = bold_font
        ws["K3"].font = bold_font
        ws["L3"].font = bold_font if (ws["L3"].value or 0) >= 0 else Font(name="Arial", size=8, bold=True,
                                                                          color="00C00000")
        ws["L3"].number_format = "#,##0"
        ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
        ws["K3"].alignment = center
        ws["L3"].alignment = right
        total_row = data_end_row
        ws.cell(row=total_row, column=2).font = bold_font
        ws.cell(row=total_row, column=2).alignment = right
        ws.cell(row=total_row, column=9).font = bold_font
        ws.cell(row=total_row, column=11).font = bold_font
        ws.cell(row=total_row, column=9).number_format = "#,##0"
        ws.cell(row=total_row, column=11).number_format = "#,##0"

    @classmethod
    def build_statement_excel(cls, customer_id, date_from=None, date_to=None):
        data = cls.build_statement(customer_id=customer_id, date_from=date_from, date_to=date_to)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["C1"] = f"{data['from'].strftime('%d.%m.%Y')} 0:00:00"
        ws["C2"] = f"{data['to'].strftime('%d.%m.%Y')} 23:59:59"
        ws["C3"] = data["customer"]["full_name"]
        ws["K3"] = "Остаток"
        ws["L3"] = cls._fmt_number(data["opening_balance"])
        ws["B4"] = "№"
        ws["C4"] = "Дата"
        ws["D4"] = "Регистратор"
        ws["E4"] = "Вид оплаты"
        ws["F4"] = "ТуловМаксади"
        ws["G4"] = "Товар"
        ws["H4"] = "Приход"
        ws["J4"] = "Расход"
        ws["L4"] = "Остаток"
        ws["H5"] = "Кол"
        ws["I5"] = "Сумма"
        ws["J5"] = "Кол"
        ws["K5"] = "Сумма"

        row_idx = 6
        data_start_row = row_idx

        for row in data["rows"]:
            ws.cell(row=row_idx, column=2, value=row["no"])
            ws.cell(row=row_idx, column=3, value=row["date"].strftime("%d.%m.%Y"))
            ws.cell(row=row_idx, column=4, value=row["registrator"])
            ws.cell(row=row_idx, column=5, value=row["payment_type"])
            ws.cell(row=row_idx, column=6, value=row["purpose"])
            ws.cell(row=row_idx, column=7, value=row["product"])
            ws.cell(row=row_idx, column=8, value=cls._fmt_number(row["income_qty"]))
            ws.cell(row=row_idx, column=9, value=cls._fmt_number(row["income_amount"]))
            ws.cell(row=row_idx, column=10, value=cls._fmt_number(row["expense_qty"]))
            ws.cell(row=row_idx, column=11, value=cls._fmt_number(row["expense_amount"]))
            ws.cell(row=row_idx, column=12, value=cls._fmt_number(row["balance"]))
            row_idx += 1

        ws.cell(row=row_idx, column=2, value="Жами:")
        ws.cell(row=row_idx, column=9, value=cls._fmt_number(data["totals"]["income_amount"]))
        ws.cell(row=row_idx, column=11, value=cls._fmt_number(data["totals"]["expense_amount"]))
        cls._apply_table_style(ws, data_start_row=data_start_row, data_end_row=row_idx)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
