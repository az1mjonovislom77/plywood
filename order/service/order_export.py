from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from decimal import Decimal


def generate_order_ledger_excel(order):
    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    number_format = "#,##0"

    def money(cell, val):
        cell.value = float(val)
        cell.number_format = number_format

    ws["A4"] = order.customer.full_name if order.customer else ""

    if order.customer:
        prev = order.customer.orders.filter(created_at__lt=order.created_at)
        previous_total = sum((o.total_price for o in prev), Decimal("0"))
        previous_paid = sum((o.covered_amount for o in prev), Decimal("0"))
    else:
        previous_total = Decimal("0")
        previous_paid = Decimal("0")

    balance = previous_total - previous_paid

    ws.merge_cells("I4:J4")
    ws["I4"] = "Остаток"
    ws["I4"].font = bold
    ws["I4"].alignment = center

    ws.merge_cells("I5:J5")
    money(ws["I5"], balance)
    ws["I5"].alignment = center

    ws.merge_cells("A6:A7")
    ws.merge_cells("B6:B7")
    ws.merge_cells("C6:C7")
    ws.merge_cells("D6:D7")
    ws.merge_cells("E6:E7")
    ws.merge_cells("F6:G6")
    ws.merge_cells("H6:I6")
    ws.merge_cells("J6:J7")

    ws["A6"] = "№"
    ws["B6"] = "Дата"
    ws["C6"] = "Регистратор"
    ws["D6"] = "Вид оплаты"
    ws["E6"] = "Товар"
    ws["F6"] = "Приход"
    ws["H6"] = "Расход"
    ws["J6"] = "Остаток"

    ws["F7"] = "Кол"
    ws["G7"] = "Сумма"
    ws["H7"] = "Кол"
    ws["I7"] = "Сумма"

    for r in range(6, 8):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center
            cell.border = border

    row = 8
    i = 1

    for item in order.items.select_related("product", "banding", "cutting"):
        qty = item.quantity
        amount = item.price * qty

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=str(order.created_at.date()))
        ws.cell(row=row, column=3, value=f"Order {order.id}")
        ws.cell(row=row, column=4, value=order.get_payment_method_display())
        ws.cell(row=row, column=5, value=item.product.name)

        ws.cell(row=row, column=8, value=float(qty))
        money(ws.cell(row=row, column=9), amount)

        balance -= amount
        money(ws.cell(row=row, column=10), balance)

        for c in range(1, 11):
            ws.cell(row=row, column=c).border = border

        row += 1
        i += 1

        if item.banding:
            b = item.banding
            total = b.length * (b.thickness.price if b.thickness else Decimal("0"))

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(order.created_at.date()))
            ws.cell(row=row, column=3, value=f"Order {order.id}")
            ws.cell(row=row, column=4, value=order.get_payment_method_display())
            ws.cell(row=row, column=5, value=f"Kromka {b.length}m")

            ws.cell(row=row, column=8, value=float(b.length))
            money(ws.cell(row=row, column=9), total)

            balance -= total
            money(ws.cell(row=row, column=10), balance)

            for c in range(1, 11):
                ws.cell(row=row, column=c).border = border

            row += 1
            i += 1

        if item.cutting:
            c = item.cutting
            total = c.count * c.price

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(order.created_at.date()))
            ws.cell(row=row, column=3, value=f"Order {order.id}")
            ws.cell(row=row, column=4, value=order.get_payment_method_display())
            ws.cell(row=row, column=5, value="Kesish")

            ws.cell(row=row, column=8, value=float(c.count))
            money(ws.cell(row=row, column=9), total)

            balance -= total
            money(ws.cell(row=row, column=10), balance)

            for c in range(1, 11):
                ws.cell(row=row, column=c).border = border

            row += 1
            i += 1

    ws.cell(row=row, column=4, value="Жами:").font = bold
    money(ws.cell(row=row, column=9), order.total_price)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer