from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from product.models import Product
from acceptance.models import CurrencyRate


class ProductExcelExportService:
    @staticmethod
    def _get_rate():
        rate_obj = CurrencyRate.objects.filter(date__lte=timezone.localdate()).order_by("-date").first()
        return rate_obj.rate if rate_obj else None

    @classmethod
    def build_excel(cls, queryset, user=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        rate = cls._get_rate()

        # Headers
        headers = [
            "ID",
            "Категория",
            "Номи",
            "Ранги (color)",
            "Сифати",
            "Узунлиги",
            "Эни",
            "Қалинлиги",
            "Келиш нархи (so'm)", 
            "Келиш нархи ($)", 
            "Инвестиция ($)",
            "Сотиш нархи (so'm)",
            "Сотиш нархи ($)",
            "Омбордаги қолдиқ (count)",
            "Қабул қилинган сана",
            "Изоҳ"
        ]

        ws.append(headers)

        # Header styles
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        total_investment = 0
        total_count = 0

        # Data rows
        for product in queryset:
            category_name = product.category.name if product.category else ""
            
            # Calculate sale price in dollar
            sale_price_in_dollar = 0
            if rate and product.sale_price:
                sale_price_in_dollar = float((product.sale_price / rate).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

            # Kelish narxi va investitsiya qiymatlarini ishonchli olish
            arrival_price = float(product.arrival_price) if product.arrival_price else 0
            arrival_price_in_dollar = float(product.arrival_price_in_dollar) if product.arrival_price_in_dollar else 0
            
            # Agar bazadagi arrival_price_in_dollar 0 bo'lsa (yoki saqlanmay qolgan bo'lsa), joriy kurs bo'yicha hisoblaymiz:
            if arrival_price_in_dollar == 0 and rate and product.arrival_price:
                arrival_price_in_dollar = float((product.arrival_price / rate).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

            count = float(product.count) if product.count else 0
            
            # Investitsiya har doim aniq hisoblanishi uchun:
            investment_in_dollar = count * arrival_price_in_dollar
            
            total_investment += investment_in_dollar
            total_count += count

            row = [
                product.id,
                category_name,
                product.name,
                product.color if product.color else "",
                product.get_quality_display() if product.quality else "",
                float(product.width) if product.width else 0,
                float(product.height) if product.height else 0,
                float(product.thick) if product.thick else 0,
                arrival_price,
                arrival_price_in_dollar,
                investment_in_dollar,
                float(product.sale_price) if product.sale_price else 0,
                sale_price_in_dollar,
                count,
                product.arrival_date.strftime('%Y-%m-%d') if product.arrival_date else "",
                product.description if product.description else ""
            ]
            
            ws.append(row)

        # Add total row
        total_row = [
            "", "", "ЖАМИ", "", "", "", "", "", "", "", total_investment, "", "", total_count, "", ""
        ]
        ws.append(total_row)
        
        last_row_idx = ws.max_row
        for cell in ws[last_row_idx]:
            cell.font = bold_font
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        ws.merge_cells(start_row=last_row_idx, start_column=1, end_row=last_row_idx, end_column=3)
        ws.cell(row=last_row_idx, column=1).alignment = Alignment(horizontal="right", vertical="center")

        # Data styles
        for row in ws.iter_rows(min_row=2, max_row=last_row_idx - 1):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)) and cell.column not in [1]: # Don't format ID
                    cell.number_format = '#,##0.00'

        # Set column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 20
        ws.column_dimensions["N"].width = 20
        ws.column_dimensions["O"].width = 15
        ws.column_dimensions["P"].width = 30

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
