from io import BytesIO
from decimal import Decimal
from django.db.models import Sum, Value, F, ExpressionWrapper, Q, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from category.models import Category
from employee.models import SalaryPayment
from product.models import Product
from acceptance.models import Acceptance, CurrencyRate
from order.models import Order, OrderItem
from product.services.material_profit import MaterialProfitService
from utils.models import Expenses


class MaterialReportService:
    @staticmethod
    def _money_field():
        return DecimalField(max_digits=18, decimal_places=2)

    @classmethod
    def _money_expr(cls, left, right):
        return ExpressionWrapper(F(left) * F(right), output_field=cls._money_field())

    @staticmethod
    def _accepted_order_filter():
        return Q(order__order_status=Order.OrderStatus.ACCEPT)

    @staticmethod
    def _accepted_order_before_filter(start_dt):
        return Q(order__accepted_at__lt=start_dt) | Q(order__accepted_at__isnull=True, order__created_at__lt=start_dt)

    @staticmethod
    def _accepted_order_range_filter(start_dt, end_dt):
        return Q(order__accepted_at__gte=start_dt, order__accepted_at__lt=end_dt) | Q(
            order__accepted_at__isnull=True,
            order__created_at__gte=start_dt,
            order__created_at__lt=end_dt,
        )

    @classmethod
    def _parse_bounds(cls, date_from, date_to):
        today = timezone.localdate()
        start_date = parse_date(date_from) if date_from else today
        end_date = parse_date(date_to) if date_to else today

        if not start_date or not end_date:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        if end_date < start_date:
            raise ValueError("to date must be greater than or equal to from date")

        start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(
            timezone.datetime.combine(end_date + timezone.timedelta(days=1), timezone.datetime.min.time()))
        return start_date, end_date, start_dt, end_dt

    @staticmethod
    def _to_map(qs, extra_fields=None):
        if extra_fields is None:
            extra_fields = []
        data = {}
        for row in qs:
            product_id = row["product_id"]
            data[product_id] = {
                "qty": Decimal(str(row.get("qty") or 0)),
                "total": Decimal(str(row.get("total") or 0)),
            }
            for field in extra_fields:
                data[product_id][field] = Decimal(str(row.get(field) or 0))
        return data

    @classmethod
    def build_excel(cls, date_from=None, date_to=None):
        profit_context = MaterialProfitService.build_profit_context(date_from, date_to)
        start_date = profit_context["start_date"]
        end_date = profit_context["end_date"]
        start_dt = profit_context["start_dt"]
        end_dt = profit_context["end_dt"]
        open_cogs_map = profit_context["open_cogs_map"]
        period_cogs_map = profit_context["period_cogs_map"]
        period_cogs_map_in_dollar = profit_context["period_cogs_map_in_dollar"]

        categories = list(Category.objects.all().order_by("name"))
        products = list(Product.objects.select_related("category")
                        .order_by("category__name", "name"))

        open_in_map = cls._to_map(
            Acceptance.objects.filter(acceptance_status="accept", arrival_date__lt=start_date)
            .values("product_id").annotate(
                qty=Coalesce(Sum("count"), Value(Decimal("0")), output_field=cls._money_field()),
                total=Coalesce(
                    Sum(cls._money_expr("count", "arrival_price")), Value(Decimal("0")),
                    output_field=cls._money_field(),
                ),
            )
        )

        open_out_map = cls._to_map(
            OrderItem.objects.filter(
                cls._accepted_order_filter(), cls._accepted_order_before_filter(start_dt))
            .values("product_id").annotate(
                qty=Coalesce(Sum("quantity"), Value(Decimal("0")), output_field=cls._money_field()),
                total=Coalesce(Sum(cls._money_expr("quantity", "price")),
                               Value(Decimal("0")), output_field=cls._money_field())))

        in_map = cls._to_map(
            Acceptance.objects.filter(
                acceptance_status="accept", arrival_date__gte=start_date, arrival_date__lte=end_date,
            ).values("product_id").annotate(
                qty=Coalesce(Sum("count"), Value(Decimal("0")), output_field=cls._money_field()),
                total=Coalesce(Sum(cls._money_expr("count", "arrival_price")),
                               Value(Decimal("0")), output_field=cls._money_field())))

        out_map = cls._to_map(
            OrderItem.objects.filter(cls._accepted_order_filter(), cls._accepted_order_range_filter(start_dt, end_dt))
            .values("product_id").annotate(
                qty=Coalesce(Sum("quantity"), Value(Decimal("0")), output_field=cls._money_field()),
            )
        )
        revenue_som_map = profit_context["revenue_som_map"]
        revenue_dollar_map = profit_context["revenue_dollar_map"]
        grouped_products = {}
        for product in products:
            grouped_products.setdefault(product.category_id, []).append(product)

        wb = Workbook()
        ws = wb.active
        ws.title = "Material Report"
        bold = Font(name="Arial", size=10, bold=True)
        normal = Font(name="Arial", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def money(cell, value):
            value = float(value or 0)
            if value == int(value):
                cell.value = int(value)
                cell.number_format = "#,##0"
            else:
                cell.value = value
                cell.number_format = "#,##0.000"

        def money_with_dollar(cell, value, value_in_dollar):
            s_value = f"{float(value or 0):,.2f}".replace(",", " ").replace(".", ",")
            s_value_in_dollar = f"{float(value_in_dollar or 0):,.2f}".replace(",", " ").replace(".", ",")
            cell.value = f"{s_value} ({s_value_in_dollar}$)"
            cell.alignment = right

        ws.merge_cells("B1:L1")
        ws["B1"] = f"Материальный отчет за {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws["B1"].font = Font(name="Arial", size=14, bold=True)
        ws["B1"].alignment = left
        ws["B3"] = "Склад"
        ws["B3"].font = Font(name="Arial", size=14, bold=True)
        ws["F3"] = "Асосий РМУ"
        ws["F3"].font = Font(name="Arial", size=18)
        ws["F3"].alignment = left
        ws.merge_cells("A4:A6")
        ws.merge_cells("B4:D6")
        ws.merge_cells("E4:E6")
        ws.merge_cells("F4:G4")
        ws.merge_cells("H4:I4")
        ws.merge_cells("J4:K4")
        ws.merge_cells("L4:M4")
        ws.merge_cells("N4:N6")
        ws["A4"] = "Код"
        ws["B4"] = "МатериалРодитель / Материал"
        ws["E4"] = "Ед.изм"
        ws["F4"] = "Сальдо на начало"
        ws["H4"] = "Приход"
        ws["J4"] = "Расход"
        ws["L4"] = "Сальдо на конец"
        ws["N4"] = "Соф фойда"
        ws["F5"] = "Количество"
        ws["G5"] = "Сумма"
        ws["H5"] = "Количество"
        ws["I5"] = "Сумма"
        ws["J5"] = "Количество"
        ws["K5"] = "Сумма"
        ws["L5"] = "Количество"
        ws["M5"] = "Сумма"

        for r in range(4, 7):
            for c in range(1, 15):
                cell = ws.cell(r, c)
                if isinstance(cell, MergedCell) and cell.coordinate != ws.cell(r, c).coordinate:
                    continue
                cell.font = bold
                cell.alignment = center
                cell.border = border

        row = 7
        grand_open_qty = Decimal("0")
        grand_open_sum = Decimal("0")
        grand_in_qty = Decimal("0")
        grand_in_sum = Decimal("0")
        grand_out_qty = Decimal("0")
        grand_out_sum = Decimal("0")
        grand_out_revenue_sum = Decimal("0")
        grand_out_revenue_sum_in_dollar = Decimal("0")
        grand_out_sum_in_dollar = Decimal("0")
        grand_end_qty = Decimal("0")
        grand_end_sum = Decimal("0")
        grand_profit_sum = Decimal("0")
        grand_profit_sum_in_dollar = Decimal("0")

        for category in categories:
            category_products = grouped_products.get(category.id, [])
            if not category_products:
                continue

            cat_open_qty = Decimal("0")
            cat_open_sum = Decimal("0")
            cat_in_qty = Decimal("0")
            cat_in_sum = Decimal("0")
            cat_out_qty = Decimal("0")
            cat_out_sum = Decimal("0")
            cat_out_revenue_sum = Decimal("0")
            cat_out_revenue_sum_in_dollar = Decimal("0")
            cat_out_sum_in_dollar = Decimal("0")
            cat_end_qty = Decimal("0")
            cat_end_sum = Decimal("0")
            cat_profit_sum = Decimal("0")
            cat_profit_sum_in_dollar = Decimal("0")

            product_rows = []

            for product in category_products:
                open_in = open_in_map.get(product.id, {"qty": Decimal("0"), "total": Decimal("0")})
                open_out = open_out_map.get(product.id, {"qty": Decimal("0"), "total": Decimal("0")})
                in_period = in_map.get(product.id, {"qty": Decimal("0"), "total": Decimal("0")})
                out_period = out_map.get(product.id, {"qty": Decimal("0"), "total": Decimal("0")})
                out_revenue = revenue_som_map.get(product.id, Decimal("0"))
                out_revenue_in_dollar = revenue_dollar_map.get(product.id, Decimal("0"))
                out_cogs = period_cogs_map.get(product.id, Decimal("0"))
                out_cogs_in_dollar = period_cogs_map_in_dollar.get(product.id, Decimal("0"))
                out_qty = out_period["qty"]
                open_qty = open_in["qty"] - open_out["qty"]
                open_sum = open_in["total"] - open_cogs_map.get(product.id, Decimal("0"))
                in_qty = in_period["qty"]
                in_sum = in_period["total"]
                end_qty = open_qty + in_qty - out_qty
                end_sum = open_sum + in_sum - out_cogs
                profit_row = MaterialProfitService.product_profit_row(product, profit_context)
                profit_som = profit_row["profit_som"]
                profit_dollar = profit_row["profit_dollar"]
                cat_open_qty += open_qty
                cat_open_sum += open_sum
                cat_in_qty += in_qty
                cat_in_sum += in_sum
                cat_out_qty += out_qty
                cat_out_sum += out_cogs
                cat_out_revenue_sum += out_revenue
                cat_out_revenue_sum_in_dollar += out_revenue_in_dollar
                cat_out_sum_in_dollar += out_cogs_in_dollar
                cat_end_qty += end_qty
                cat_end_sum += end_sum
                cat_profit_sum += profit_som
                cat_profit_sum_in_dollar += profit_dollar

                product_rows.append({
                    "code": product.id,
                    "name": product.name,
                    "unit": getattr(product, "unit", "дона"),
                    "open_qty": open_qty,
                    "open_sum": open_sum,
                    "in_qty": in_qty,
                    "in_sum": in_sum,
                    "out_qty": out_qty,
                    "out_revenue": out_revenue,
                    "out_revenue_in_dollar": out_revenue_in_dollar,
                    "out_sum": out_cogs,
                    "out_sum_in_dollar": out_cogs_in_dollar,
                    "profit_som": profit_som,
                    "profit_dollar": profit_dollar,
                    "end_qty": end_qty,
                    "end_sum": end_sum,
                })

            grand_open_qty += cat_open_qty
            grand_open_sum += cat_open_sum
            grand_in_qty += cat_in_qty
            grand_in_sum += cat_in_sum
            grand_out_qty += cat_out_qty
            grand_out_sum += cat_out_sum
            grand_out_revenue_sum += cat_out_revenue_sum
            grand_out_revenue_sum_in_dollar += cat_out_revenue_sum_in_dollar
            grand_out_sum_in_dollar += cat_out_sum_in_dollar
            grand_end_qty += cat_end_qty
            grand_end_sum += cat_end_sum
            grand_profit_sum += cat_profit_sum
            grand_profit_sum_in_dollar += cat_profit_sum_in_dollar
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row, 2, category.name)
            ws.cell(row, 2).font = bold
            ws.cell(row, 2).alignment = left
            money(ws.cell(row, 6), cat_open_qty)
            money(ws.cell(row, 7), cat_open_sum)
            money(ws.cell(row, 8), cat_in_qty)
            money(ws.cell(row, 9), cat_in_sum)
            money(ws.cell(row, 10), cat_out_qty)
            money_with_dollar(ws.cell(row, 11), cat_out_sum, cat_out_revenue_sum_in_dollar)
            money(ws.cell(row, 12), cat_end_qty)
            money(ws.cell(row, 13), cat_end_sum)
            money_with_dollar(ws.cell(row, 14), cat_profit_sum, cat_profit_sum_in_dollar)

            for c in range(1, 15):
                cell = ws.cell(row, c)
                if isinstance(cell, MergedCell):
                    continue
                cell.border = border
                cell.font = bold

            row += 1

            for item in product_rows:
                ws.cell(row, 1, item["code"])
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                ws.cell(row, 2, item["name"])
                ws.cell(row, 5, item["unit"])
                money(ws.cell(row, 6), item["open_qty"])
                money(ws.cell(row, 7), item["open_sum"])
                money(ws.cell(row, 8), item["in_qty"])
                money(ws.cell(row, 9), item["in_sum"])
                money(ws.cell(row, 10), item["out_qty"])
                money_with_dollar(ws.cell(row, 11), item.get("out_sum", Decimal("0")),
                                  item.get("out_revenue_in_dollar", Decimal("0")))
                money(ws.cell(row, 12), item["end_qty"])
                money(ws.cell(row, 13), item["end_sum"])
                money_with_dollar(ws.cell(row, 14), item.get("profit_som", Decimal("0")),
                                  item.get("profit_dollar", Decimal("0")))

                for c in range(1, 15):
                    cell = ws.cell(row, c)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.border = border
                    cell.font = normal
                    cell.alignment = left if c in [1, 2, 5] else right

                row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, "Жами:")
        ws.cell(row, 1).font = bold
        ws.cell(row, 1).alignment = right
        money(ws.cell(row, 6), grand_open_qty)
        money(ws.cell(row, 7), grand_open_sum)
        money(ws.cell(row, 8), grand_in_qty)
        money(ws.cell(row, 9), grand_in_sum)
        money(ws.cell(row, 10), grand_out_qty)
        money_with_dollar(ws.cell(row, 11), grand_out_sum, grand_out_revenue_sum_in_dollar)
        money(ws.cell(row, 12), grand_end_qty)
        money(ws.cell(row, 13), grand_end_sum)
        money_with_dollar(ws.cell(row, 14), grand_profit_sum, grand_profit_sum_in_dollar)

        for c in range(1, 15):
            cell = ws.cell(row, c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = border
            cell.font = bold

        widths = {"A": 12, "B": 42, "C": 2, "D": 2, "E": 25, "F": 12, "G": 18, "H": 12, "I": 18, "J": 25, "K": 18,
                  "L": 12, "M": 18, "N": 18}

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        salary_total = (
                SalaryPayment.objects.filter(paid_at__gte=start_dt, paid_at__lt=end_dt)
                .aggregate(
                    total=Coalesce(Sum("amount"),
                                   Value(Decimal("0")), output_field=cls._money_field()))["total"] or Decimal("0"))

        row += 4
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        title_cell = ws.cell(row, 2)
        title_cell.value = "ФОЙДА ВА ХАРАЖАТЛАР"
        title_cell.font = Font(name="Arial", size=16, bold=True)
        title_cell.alignment = center

        row += 2

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        profit_header = ws.cell(row, 2)
        profit_header.value = "ФОЙДА"
        profit_header.font = bold
        profit_header.alignment = center

        for col in range(2, 6):
            ws.cell(row, col).border = border

        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        expense_header = ws.cell(row, 7)
        expense_header.value = "ХАРАЖАТ"
        expense_header.font = bold
        expense_header.alignment = center

        for col in range(7, 11):
            ws.cell(row, col).border = border

        row += 1

        ws.cell(row, 2, "НОМИ")
        ws.cell(row, 5, "СУММА")
        ws.cell(row, 7, "НОМИ")
        ws.cell(row, 10, "СУММА")

        for col in [2, 5, 7, 10]:
            ws.cell(row, col).font = bold
            ws.cell(row, col).alignment = center
            ws.cell(row, col).border = border

        row += 1

        profit_rows = []

        for category in categories:
            category_products = grouped_products.get(category.id, [])

            if not category_products:
                continue

            category_profit = Decimal("0")

            for product in category_products:
                profit_row = MaterialProfitService.product_profit_row(product, profit_context)
                category_profit += profit_row["profit_som"]
            profit_rows.append((category.name, category_profit))
        profit_total = Decimal("0")

        for idx, (name, amount) in enumerate(profit_rows):
            current_row = row + idx
            ws.cell(current_row, 2, name)
            money(ws.cell(current_row, 5), amount)
            ws.cell(current_row, 2).border = border
            ws.cell(current_row, 5).border = border
            profit_total += amount

        salary_total = (
                SalaryPayment.objects.filter(paid_at__gte=start_dt, paid_at__lt=end_dt)
                .aggregate(
                    total=Coalesce(Sum("amount"), Value(Decimal("0")), output_field=cls._money_field())
                )["total"] or Decimal("0")
        )

        expense_total = (
                Expenses.objects.filter(
                    created_at__gte=start_dt, created_at__lt=end_dt,
                    expense_status__in=[Expenses.ExpensesStatus.CREATED, Expenses.ExpensesStatus.ACCEPT])
                .aggregate(
                    total=Coalesce(
                        Sum("value"), Value(Decimal("0")), output_field=cls._money_field()))["total"] or Decimal("0")
        )

        expense_rows = [("Ходимлар ойлиги", salary_total), ("Бошқа харажатлар", expense_total)]
        expense_total_sum = Decimal("0")

        for idx, (name, amount) in enumerate(expense_rows):
            current_row = row + idx
            ws.cell(current_row, 7, name)
            money(ws.cell(current_row, 10), amount)
            ws.cell(current_row, 7).border = border
            ws.cell(current_row, 10).border = border
            expense_total_sum += amount

        rate_obj = CurrencyRate.objects.filter(
            date=timezone.localdate()
        ).first()

        rate_value = Decimal(str(rate_obj.rate)) if rate_obj else Decimal("0")

        profit_dollar = Decimal("0")
        expense_dollar = Decimal("0")

        if rate_value:
            profit_dollar = profit_total / rate_value
            expense_dollar = expense_total_sum / rate_value

        summary_row = row + max(
            len(profit_rows),
            len(expense_rows),
        )

        ws.cell(summary_row, 2, "ЖАМИ ФОЙДА")

        money_with_dollar(
            ws.cell(summary_row, 5),
            profit_total,
            profit_dollar,
        )

        ws.cell(summary_row, 7, "ЖАМИ ХАРАЖАТ")

        money_with_dollar(
            ws.cell(summary_row, 10),
            expense_total_sum,
            expense_dollar,
        )

        for col in [2, 5, 7, 10]:
            ws.cell(summary_row, col).font = bold
            ws.cell(summary_row, col).border = border
        summary_row += 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
