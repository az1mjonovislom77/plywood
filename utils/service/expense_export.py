from io import BytesIO
from decimal import Decimal
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from customer.models import BalanceHistory
from employee.models import SalaryPayment
from order.models import Order, Banding, Cutting
from supplier.models import SupplierTransaction
from utils.models import Expenses, Services
from utils.service.comprehensive_stats import DashboardStatsService


class CashFlowReportService:
    @classmethod
    def _parse_bounds(cls, date_from, date_to):
        today = timezone.localdate()
        start_date = parse_date(date_from) if date_from else today
        end_date = parse_date(date_to) if date_to else today

        if not start_date or not end_date:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        if end_date < start_date:
            raise ValueError("to date must be greater than or equal to from date")

        start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(
            timezone.datetime.combine(end_date + timezone.timedelta(days=1), timezone.datetime.min.time()))
        return start_date, end_date, start_dt, end_dt

    @classmethod
    def build_excel(cls, date_from=None, date_to=None):
        start_date, end_date, start_dt, end_dt = cls._parse_bounds(date_from, date_to)
        income_rows = []
        expense_rows = []

        for order in Order.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt, order_status=Order.OrderStatus.ACCEPT,
                covered_amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            income_rows.append({
                "name": order.customer.full_name if order.customer else "Аноним",
                "amount": Decimal(str(order.covered_amount)),
                "created_at": order.created_at,
            })

        for banding in Banding.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt, covered_amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            income_rows.append({
                "name": banding.customer.full_name if banding.customer else "Аноним",
                "amount": Decimal(str(banding.covered_amount)),
                "created_at": banding.created_at,
            })

        for cutting in Cutting.objects.filter(
                created_at__gte=start_dt,
                created_at__lt=end_dt,
                covered_amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            income_rows.append({
                "name": cutting.customer.full_name if cutting.customer else "Аноним",
                "amount": Decimal(str(cutting.covered_amount)),
                "created_at": cutting.created_at,
            })

        for service in Services.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt, covered_amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            income_rows.append({
                "name": service.customer.full_name if service.customer else "Аноним",
                "amount": Decimal(str(service.covered_amount)),
                "created_at": service.created_at,
            })

        for payment in BalanceHistory.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt, type=BalanceHistory.Type.PAYMENT, amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            income_rows.append({
                "name": payment.customer.full_name if payment.customer else "Аноним",
                "amount": Decimal(str(payment.amount)),
                "created_at": payment.created_at,
            })

        income_rows.sort(key=lambda x: (x["created_at"], x["name"]))

        for exp in Expenses.objects.filter(
                created_at__gte=start_dt,
                created_at__lt=end_dt,
                expense_status__in=[
                    Expenses.ExpensesStatus.ACCEPT,
                    Expenses.ExpensesStatus.CREATED,
                ],
        ).order_by("created_at", "id"):
            expense_rows.append({
                "description": exp.description,
                "created_at": exp.created_at,
                "amount": Decimal(str(exp.value)),
            })

        for payment in SupplierTransaction.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt,
                transaction_type=SupplierTransaction.TransactionType.PAYMENT, amount__gt=0,
        ).select_related("supplier").order_by("created_at", "id"):
            expense_rows.append({
                "description": payment.supplier.full_name if payment.supplier else "Поставщик",
                "created_at": payment.created_at,
                "amount": Decimal(str(payment.amount)),
            })

        for payment in (SalaryPayment.objects.filter(paid_at__gte=start_dt, paid_at__lt=end_dt)
                .select_related("employee").order_by("paid_at", "id")):
            expense_rows.append({
                "description": f"{payment.employee.full_name} (Ходим)",
                "created_at": payment.paid_at,
                "amount": Decimal(str(payment.amount)),
            })

        for refund in BalanceHistory.objects.filter(
                created_at__gte=start_dt, created_at__lt=end_dt, type=BalanceHistory.Type.REFUND, amount__gt=0,
        ).select_related("customer").order_by("created_at", "id"):
            c_name = refund.customer.full_name if refund.customer else "Аноним"
            expense_rows.append({
                "description": f"Qaytarish - {c_name}",
                "created_at": refund.created_at,
                "amount": Decimal(str(refund.amount)),
            })

        expense_rows.sort(key=lambda x: (x["created_at"], x["description"]))
        opening_balance = Decimal(str(DashboardStatsService._cashbox_total(end_dt=start_dt)))
        income_total = sum((row["amount"] for row in income_rows), Decimal("0"))
        expense_total = sum((row["amount"] for row in expense_rows), Decimal("0"))
        closing_balance = Decimal(str(DashboardStatsService._cashbox_total(end_dt=end_dt)))
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow"
        bold = Font(name="Arial", size=12, bold=True)
        normal = Font(name="Arial", size=12)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def money(cell, value):
            cell.value = float(value or 0)
            cell.number_format = "#,##0"

        ws.merge_cells("B1:H1")
        ws[
            "B1"] = f"Отчет по движению денежных средств за {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws["B1"].font = bold
        ws["B1"].alignment = left
        ws["D2"] = "Остаток на начало периода :"
        ws["D3"] = "Остаток на конец периода :"
        ws["E2"] = float(opening_balance)
        ws["E3"] = float(closing_balance)

        for c in ["D2", "D3", "E2", "E3"]:
            ws[c].font = bold

        ws["D2"].alignment = right
        ws["D3"].alignment = right
        ws["E2"].alignment = left
        ws["E3"].alignment = left
        ws["E2"].number_format = "#,##0"
        ws["E3"].number_format = "#,##0"
        ws["B5"] = "№"
        ws["C5"] = "Контрагент номи"
        ws["D5"] = "Приход"
        ws["F5"] = "Izoh"
        ws["G5"] = "Сана"
        ws["H5"] = "Расход"

        for cell in ["B5", "C5", "D5", "F5", "G5", "H5"]:
            ws[cell].font = bold
            ws[cell].alignment = center
            ws[cell].border = border

        left_row = 6
        right_row = 6

        for idx, row in enumerate(income_rows, start=1):
            ws.cell(left_row, 2, idx)
            ws.cell(left_row, 3, row["name"])
            money(ws.cell(left_row, 4), row["amount"])
            ws.cell(left_row, 2).font = normal
            ws.cell(left_row, 3).font = normal
            ws.cell(left_row, 4).font = normal
            ws.cell(left_row, 2).alignment = center
            ws.cell(left_row, 3).alignment = left
            ws.cell(left_row, 4).alignment = right

            for col in range(2, 5):
                ws.cell(left_row, col).border = border

            left_row += 1

        for row in expense_rows:
            ws.cell(right_row, 6, row["description"])
            ws.cell(right_row, 7, row["created_at"].strftime("%d.%m.%Y"))
            money(ws.cell(right_row, 8), row["amount"])
            ws.cell(right_row, 6).font = normal
            ws.cell(right_row, 7).font = normal
            ws.cell(right_row, 8).font = normal
            ws.cell(right_row, 6).alignment = left
            ws.cell(right_row, 7).alignment = center
            ws.cell(right_row, 8).alignment = right

            for col in range(6, 9):
                ws.cell(right_row, col).border = border

            right_row += 1

        ws.merge_cells(f"B{left_row}:C{left_row}")
        ws[f"B{left_row}"] = "Жами:"
        ws[f"B{left_row}"].font = bold
        ws[f"B{left_row}"].alignment = right
        money(ws[f"D{left_row}"], income_total)
        ws[f"D{left_row}"].font = bold

        for col in range(2, 5):
            ws.cell(left_row, col).border = border

        ws.merge_cells(f"F{right_row}:G{right_row}")
        ws[f"F{right_row}"] = "Жами:"
        ws[f"F{right_row}"].font = bold
        ws[f"F{right_row}"].alignment = right
        money(ws[f"H{right_row}"], expense_total)
        ws[f"H{right_row}"].font = bold

        for col in range(6, 9):
            ws.cell(right_row, col).border = border

        widths = {
            "B": 6,
            "C": 34,
            "D": 16,
            "E": 4,
            "F": 34,
            "G": 14,
            "H": 16,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
